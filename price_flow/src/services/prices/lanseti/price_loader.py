import asyncio
import binascii
import email
import imaplib
import os
import re

from collections.abc import Callable
from datetime import UTC, datetime
from email.message import Message
from pathlib import Path
from typing import Annotated, Any, ClassVar

import pandas as pd
import requests

from bs4 import BeautifulSoup
from fastapi import Depends
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from core.exceptions.app_exceptions import (
    DriveApiError,
    EmailFetchError,
    ExcelProcessingError,
    PriceProcessingError,
    SupplierDataError,
)
from core.exceptions.base import BaseAppException
from core.exceptions.file import FileAppNotFoundError
from core.logger import logger
from core.settings import settings
from models.supplier_models import SupplierProductCode
from repositories.supplier_product_codes_repo import (
    SupplierProductCodeRepository,
    get_supplier_product_codes_repo,
)
from schemas.converter_schemas import UploadResult
from services.converter import FileUploader, get_file_uploader


class PriceLoader:
    """
    Сервис для загрузки и обработки прайс-листов Ланцети.
    Поддерживает загрузку из Gmail, поиск в Google Drive и обогащение данных.
    """

    # ===== Константы =====
    EMAIL_SCAN_LIMIT: ClassVar[int] = 200
    IMAP_HOST: ClassVar[str] = "imap.gmail.com"
    LINK_TRACKER_SUBSTR: ClassVar[str] = "geteml.com/ru/mail_link_tracker"
    DEFAULT_SUPPLIER_ID: ClassVar[int] = 201
    TARGET_FILENAME: ClassVar[tuple[str, ...]] = (
        "Нал основной прайс на элитку BY",
        "Нал основной прайс на элитку KZ",
        "Нал миниатюры дезодоранты тестеры основной прайс на элитку",
    )

    # Правила для автоматического заполнения
    PRODUCT_NAME_RULES: ClassVar[
        list[tuple[Callable[[str], bool], str, str]]
    ] = [
        # (условие, группа, подгруппа)
        (
            lambda name: any(
                keyword in name.lower() for keyword in ["лицензия", "***"]
            ),
            "Элит Парфюм",
            "Лицензия***",
        ),
        (
            lambda name: any(
                keyword in name.lower()
                for keyword in [
                    "vintag",
                    "винтаж",
                    "novaya zarya",
                    "косметика",
                    "новая заря",
                ]
            ),
            "NO",
            "NO",
        ),
        (
            lambda name: "montale" in name.lower()
            and "декодированный" in name.lower(),
            "NO",
            "NO",
        ),
    ]

    def __init__(
        self,
        settings: Any,
        supplier_codes_repo: SupplierProductCodeRepository,
        file_uploader: FileUploader,
        supplier_id: int = DEFAULT_SUPPLIER_ID,
        target_filename: tuple[str, ...] = TARGET_FILENAME,
    ) -> None:
        """
        Инициализация сервиса.

        Args:
            settings: Объект настроек приложения
            supplier_codes_repo: Репозиторий для работы с данными поставщика
            supplier_id: ID поставщика для обработки
        """
        self.settings = settings
        self.supplier_codes_repo = supplier_codes_repo
        self.file_uploader = file_uploader
        self.supplier_id = supplier_id
        self.target_filename = target_filename
        self._drive_service = None

        logger.info(
            "PriceLoader initialized",
            extra={"supplier_id": supplier_id},
        )

    # ----- Публичный метод -----
    async def process_price(
        self,
        output_filename: str | None = None,
    ) -> UploadResult:
        """
        Основной метод:
        получение ссылки -> поиск файла -> скачивание -> обработка.

        Returns:
            UploadResult: результат загрузки обработанного файла.
        """
        logger.info("Starting price processing")
        start_time = datetime.now(UTC)
        enriched_file_path: Path | None = None

        try:
            # 1. Получение ссылки из почты
            logger.info("Fetching drive link from email...")
            drive_link = await self._get_latest_drive_link()
            self._validate_drive_link(drive_link)
            logger.info(
                f"Find link on Google Drive: {drive_link[:50]}..."  # type: ignore[index]
            )

            # 2. Поиск файла в Google Drive
            logger.info(f"Searching for file '{self.target_filename[0]}'...")
            file_id = await self._find_file_in_drive(
                drive_link,  # type: ignore[arg-type]
                self.target_filename,
            )
            self._validate_file_id(file_id, self.target_filename)  # type: ignore[arg-type]
            logger.info(f"Find file with ID: {file_id}")

            # 3. Скачивание файла
            if not output_filename:
                timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
                output_filename = f"price_{self.supplier_id}_{timestamp}.xlsx"

            output_path = settings.app.base_dir / Path(
                f"uploads/{output_filename}"
            )
            output_path.parent.mkdir(parents=True, exist_ok=True)

            logger.info(f"Downloading file to {output_path}")
            await self._download_file(file_id, output_path)  # type: ignore[arg-type]

            # 4. Получение данных поставщика
            logger.info(f"Loading supplier data (ID={self.supplier_id})...")
            supplier_df = await self._get_supplier_data_async()

            # 5. Обработка Excel файла
            logger.info("Processing Excel file...")
            enriched_file_path = await self._process_excel_file(
                output_path, supplier_df
            )

            # 6. Конвертация
            logger.info("Uploading enriched file...")
            upload_result = self.file_uploader.upload_file(enriched_file_path)

            # 7. Статистика
            elapsed = (datetime.now(UTC) - start_time).total_seconds()
            logger.info(
                "Price processing completed",
                extra={
                    "time_seconds": elapsed,
                    "result_path": str(enriched_file_path),
                },
            )

        except BaseAppException:
            raise
        except Exception as e:
            logger.error("Unexpected error", extra={"error": str(e)})
            # Оборачиваем неизвестные исключения
            raise PriceProcessingError(
                error_code="PRICE_PROCESSING_ERROR",
                message="Unexpected error during price processing",
                details=str(e),
            ) from e
        else:
            return upload_result
        finally:
            if enriched_file_path and enriched_file_path.exists():
                await self._cleanup_temp_files([enriched_file_path])

    # ----- Работа с почтой -----
    async def _get_latest_drive_link(self) -> str | None:
        """
        Получает ссылку на папку Google Drive из последнего письма.

        Returns:
            str: Ссылка на Google Drive или None если не найдена
        """
        try:
            link = await asyncio.to_thread(self._fetch_link_from_email)

            if link:
                # Очистка ссылки от трекеров
                clean_link = self._clean_tracker_url(link)
                logger.debug(
                    "Cleaned drive link",
                    extra={"link": clean_link[:100] + "..."},
                )
                return clean_link

        except imaplib.IMAP4.error as e:
            raise EmailFetchError(
                error_code="IMAP4_ERROR",
                message="IMAP connection failed",
                details=str(e),
            ) from e
        except Exception as e:
            raise EmailFetchError(
                error_code="EMAIL_FETCH_ERROR",
                message="Failed to fetch email",
                details=str(e),
            ) from e
        return None

    def _fetch_link_from_email(self) -> str | None:
        """
        Синхронный метод поиска ссылки в почте через IMAP.
        """
        try:
            mail = imaplib.IMAP4_SSL(self.IMAP_HOST)
            mail.login(
                self.settings.email.user_gmail,
                self.settings.email.pass_gmail,
            )

            mail.select("INBOX")

            # Получаем общее количество сообщений
            _, messages_data = mail.search(None, "ALL")
            if not messages_data[0]:
                logger.warning("В почтовом ящике нет сообщений")
                return None

            message_ids = messages_data[0].split()
            total_messages = len(message_ids)
            scan_limit = min(self.EMAIL_SCAN_LIMIT, total_messages)

            logger.info(
                f"Scanning last {scan_limit} of {total_messages} messages"
            )

            # Ищем с конца (последние письма)
            for i in range(total_messages, total_messages - scan_limit, -1):
                message_id = message_ids[i - 1].decode()

                # Получаем полное письмо
                _, msg_data = mail.fetch(message_id, "(RFC822)")

                for response_part in msg_data:
                    if isinstance(response_part, tuple):
                        msg = email.message_from_bytes(response_part[1])
                        sender = msg.get("From", "")

                        # Проверяем отправителя
                        if self.settings.price.lanset_price_sender in sender:
                            logger.debug("Found message from expected sender")

                            # Извлекаем ссылку
                            link = self._extract_link_from_email_body(msg)
                            if link:
                                logger.info("Link found")
                                mail.close()
                                mail.logout()
                                return link
            mail.close()
            mail.logout()
            logger.warning("No drive link found in scanned messages")

        except Exception as e:
            logger.error("Email fetch error", extra={"error": str(e)})
            raise
        else:
            return None

    def _extract_link_from_email_body(self, msg: Message) -> str | None:
        """
        Извлекает ссылку из тела email.

        Args:
            msg: Объект email сообщения

        Returns:
            str: Найденная ссылка или None
        """
        try:
            body = self._get_email_body(msg)
            if not body:
                return None

            soup = BeautifulSoup(body, "html.parser")

            # Ищем все ссылки содержащие трекер
            links = [
                a["href"]
                for a in soup.find_all("a", href=True)
                if self.LINK_TRACKER_SUBSTR in a["href"]
            ]

            if links:
                logger.debug(f"Find {len(links)} links with tracker")
                return links[0]  # type: ignore

            # Дополнительный поиск ссылок на Google Drive
            drive_links = [
                a["href"]
                for a in soup.find_all("a", href=True)
                if "drive.google.com" in a["href"]
            ]

            if drive_links:
                logger.debug(f"Find {len(drive_links)} links on Google Drive")
                return drive_links[0]  # type: ignore

        except (AttributeError, KeyError, TypeError, ValueError) as e:
            # Ошибки парсинга HTML или доступа к атрибутам
            logger.warning(
                "Failed to extract link from email",
                extra={"error": str(e)},
            )
            return None
        # except Exception as e:
        #     # Ловим только неизвестные исключения и логируем, но не прячем
        #     logger.error(
        #         f"Неожиданная ошибка при извлечении ссылки из письма: "
        #         f"{type(e).__name__}: {e}"
        #     )
        #     # Можно рейзить дальше или вернуть None в зависимости от
        #     # требований
        #     return None
        return None

    def _get_email_body(self, msg: Message) -> str | None:
        """Возвращает текстовое тело письма (HTML)."""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))

                # Пропускаем вложения
                if "attachment" in content_disposition:
                    continue

                if content_type == "text/html":
                    try:
                        body_payload = part.get_payload(decode=True)
                        if body_payload and isinstance(body_payload, bytes):
                            return body_payload.decode(
                                "utf-8", errors="ignore"
                            )
                        elif body_payload and isinstance(body_payload, str):
                            # Если уже строка, просто возвращаем
                            return body_payload
                    except UnicodeDecodeError:
                        # Логируем ошибку декодирования и идем к следующей
                        # части
                        logger.debug("Couldn't decode HTML part as UTF-8")
                        continue
                    except (AttributeError, TypeError, ValueError) as e:
                        # Если payload пустой или имеет неожиданный тип
                        # (например, None)
                        logger.debug(
                            "Payload structure error when reading the body "
                            f"letter: {e}"
                        )
                        continue
                    # except Exception as e:
                    #     logger.warning(
                    #         "Неожиданная ошибка при обработке части "
                    #         f"письма: {type(e).__name__}: {e}"
                    #     )
                    #     continue
        else:
            try:
                body_payload = msg.get_payload(decode=True)
                if body_payload and isinstance(body_payload, bytes):
                    return body_payload.decode("utf-8", errors="ignore")
                elif body_payload and isinstance(body_payload, str):
                    # Если уже строка, просто возвращаем
                    return body_payload
            except (
                UnicodeDecodeError,
                LookupError,
                TypeError,
                ValueError,
            ) as e:
                logger.warning(
                    f"Error decoding the email body: {type(e).__name__}: {e}"
                )
                return None
            # except Exception as e:
            #     logger.error(
            #         f"Критическая ошибка при получении тела письма: "
            #         f"{type(e).__name__}: {e}",
            #         exc_info=True
            #     )
            #     return None
        return None

    def _clean_tracker_url(self, url: str) -> str:
        """
        Очищает URL от email трекеров.

        Args:
            url: URL с трекером

        Returns:
            str: Очищенный URL
        """
        try:
            if "geteml.com" in url or "mail_link_tracker" in url:
                # Пробуем извлечь оригинальный URL из параметров
                match = re.search(r"url=([^&]+)", url)
                if match:
                    import base64

                    encoded_url = match.group(1)
                    try:
                        decoded = base64.b64decode(encoded_url).decode(
                            "utf-8"
                        )
                        logger.debug(
                            "The URL from the tracker has been decoded: "
                            f"{url}"
                        )
                    except (binascii.Error, UnicodeDecodeError) as e:
                        logger.debug(
                            f"Error decoding URL from tracker: {e}, "
                            f"URL: {url}"
                        )
                    else:
                        return decoded
        except (AttributeError, ValueError, KeyError, IndexError) as e:
            logger.warning(f"Couldn't clear URL from tracker: {e}")
        return url

    # ----- Google Drive -----
    async def _find_file_in_drive(
        self, folder_link: str, filenames: tuple[str, ...]
    ) -> str | None:
        """
        Ищет файл в папке Google Drive.

        Args:
            folder_link: Ссылка на папку Google Drive
            filenames: Имя искомого файла

        Returns:
            str: ID файла или None если не найден
        """
        for filename in filenames:
            try:
                return await asyncio.to_thread(
                    self._find_file_in_drive_sync, folder_link, filename
                )

            except HttpError as e:
                if e.resp.status == 429:
                    error_code = "DRIVE_RATE_LIMIT"
                    message = "Google Drive API rate limit exceeded"
                    details = "Retry later"
                    error = e
                error_code = "DRIVE_API_ERROR"
                message = "Google Drive API error"
                details = str(e)
                error = e
            except Exception as e:  # noqa: BLE001
                error_code = "DRIVE_UNKNOWN_ERROR"
                message = "Unexpected error while searching Drive"
                details = str(e)
                error = e
        raise DriveApiError(
            error_code=error_code,
            message=message,
            details=details,
        ) from error

    def _find_file_in_drive_sync(
        self, folder_link: str, filename: str
    ) -> str | None:
        """
        Синхронный метод поиска файла в Google Drive.
        """
        try:
            # First variant:
            # response = requests.head(link, allow_redirects=True)
            # response.raise_for_status()
            # final_url = response.url
            # Извлекаем ID папки из URL
            # folder_id = final_url.split('/')[-1].split('?')[0]

            folder_id = self._extract_folder_id(folder_link)
            if not folder_id:
                logger.error("Cannot extract folder ID from link")
                return None

            logger.debug(f"Find file '{filename}' in folder ID: {folder_id}")

            # Создаем сервис Google Drive
            service = build(
                "drive",
                "v3",
                developerKey=self.settings.email.api_key_google,
                cache_discovery=False,
            )

            # Ищем файлы в папке
            query = f"'{folder_id}' in parents and trashed = false"
            results = (
                service.files()
                .list(
                    q=query,
                    pageSize=100,
                    fields="files(id, name, mimeType)",
                    supportsAllDrives=True,
                    includeItemsFromAllDrives=True,
                )
                .execute()
            )

            files = results.get("files", [])
            logger.debug(f"Found {len(files)} files in folder")

            # Ищем файл по имени (частичное совпадение)
            filename_lower = filename.lower()
            for file in files:
                if filename_lower in file["name"].lower():
                    logger.info(
                        "File found",
                        extra={"file_name": file["name"], "id": file["id"]},
                    )
                    return file["id"]  # type: ignore[no-any-return]

            # Если точное совпадение не найдено, логируем список файлов
            logger.warning(f"File '{filename}' not found. Awailable files:")
            for file in files:
                logger.warning(f"  - {file['name']}")
        except Exception as e:
            logger.error(
                f"Error when searching for a file in Google Drive: {e}"
            )
            raise
        else:
            return None

    def _extract_folder_id(self, url: str) -> str | None:
        """
        Извлекает ID папки из URL Google Drive.

        Args:
            url: URL папки Google Drive

        Returns:
            str: ID папки или None
        """
        patterns = [
            r"/folders/([a-zA-Z0-9_-]+)",
            r"id=([a-zA-Z0-9_-]+)",
            r"/drive/folders/([a-zA-Z0-9_-]+)",
        ]

        for pattern in patterns:
            match = re.search(pattern, url)
            if match:
                folder_id = match.group(1)
                # Удаляем возможные дополнительные параметры
                if "?" in folder_id:
                    folder_id = folder_id.split("?")[0]
                return folder_id

        # Если паттерны не сработали, берем последнюю часть URL
        parts = url.strip("/").split("/")
        if parts:
            last_part = parts[-1]
            if "?" in last_part:
                last_part = last_part.split("?")[0]
            return last_part

        return None

    # ----- Вспомогательные валидации -----

    def _validate_drive_link(self, drive_link: str | None) -> None:
        if not drive_link:
            raise EmailFetchError(
                error_code="DRIVE_LINK_NOT_FOUND",
                message="Drive link not found in email",
            )

    def _validate_file_id(
        self, file_id: str | None, target_filename: str
    ) -> None:
        if not file_id:
            raise FileAppNotFoundError(
                # error_code="GOOGLE_DRIVE_FILE_NOT_FOUND",
                message=f"File '{target_filename}' not found in folder",
                path=target_filename,
            )

    # ----- Скачивание файла -----

    async def _download_file(self, file_id: str, output_path: Path) -> None:
        """
        Скачивает файл с Google Drive.

        Args:
            file_id: ID файла в Google Drive
            output_path: Путь для сохранения файла
        """
        try:
            await asyncio.to_thread(
                self._download_file_sync, file_id, output_path
            )
            logger.info("File downloaded", extra={"path": str(output_path)})
        except Exception as e:
            raise DriveApiError(
                error_code="DOWNLOAD_ERROR",
                message="Failed to download file",
                details=str(e),
            ) from e

    def _download_file_sync(self, file_id: str, output_path: Path) -> None:
        """
        Синхронный метод скачивания файла.
        """
        download_url = (
            f"https://drive.google.com/uc?id={file_id}&export=download"
        )

        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            )
        }

        try:
            response = requests.get(
                download_url, headers=headers, stream=True, timeout=30
            )
            response.raise_for_status()

            # Проверяем размер файла
            total_size = int(response.headers.get("content-length", 0))

            with Path.open(output_path, "wb") as file:
                if total_size == 0:
                    file.write(response.content)
                else:
                    downloaded = 0
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            file.write(chunk)
                            downloaded += len(chunk)

            # Проверяем, что файл скачан
            if output_path.exists() and output_path.stat().st_size > 0:
                file_size_mb = output_path.stat().st_size / (1024 * 1024)
                logger.info(
                    f"Файл скачан успешно. Размер: {file_size_mb:.2f} MB"
                )
            else:
                raise DriveApiError(
                    error_code="FILE_EMPTY_OR_NOT_FOUND",
                    message="The downloaded file is empty or does not exist",
                )
        except requests.exceptions.RequestException as e:
            logger.error(f"Network error when downloading a file: {e}")
            raise

    # ----- Получение данных поставщика -----

    async def _get_supplier_data_async(self) -> pd.DataFrame:
        """
        Асинхронно получает данные поставщика.

        Returns:
            DataFrame с данными поставщика.
        """
        try:
            supplier_data = await self.supplier_codes_repo.get_supplier_data(
                self.supplier_id
            )
            self._validate_supplier_data(supplier_data)

            logger.info(
                "Supplier data loaded",
                extra={"count": len(supplier_data)},
            )

        except SupplierDataError:
            raise
        except Exception as e:
            raise SupplierDataError(
                error_code="SUPPLIER_DATA_LOAD_ERROR",
                message="Failed to load supplier data",
                details=str(e),
            ) from e
        else:
            return pd.DataFrame(
                [
                    {
                        "code": item.code,
                        "category": item.category,
                        "subcategory": item.subcategory,
                    }
                    for item in supplier_data
                ]
            )

    def _validate_supplier_data(
        self, supplier_data: list[SupplierProductCode]
    ) -> None:
        if not supplier_data:
            raise SupplierDataError(
                error_code="EMPTY_SUPPLIER_DATA",
                message=f"Нет данных для поставщика ID={self.supplier_id}",
                details="Проверьте наличие записей в базе данных",
            )

    # ----- Обработка Excel -----

    async def _process_excel_file(
        self, file_path: Path, supplier_data: pd.DataFrame
    ) -> Path:
        """
        Обрабатывает Excel файл: чтение, объединение, обогащение данных.

        Args:
            file_path: Путь к Excel файлу
            supplier_data: list[SupplierProduct] с данными поставщика

        Returns:
            Path: Путь к обработанному файлу
        """
        try:
            # 1. Чтение Excel файла
            logger.info("Reading Excel file", extra={"path": str(file_path)})
            excel_df, header_row = await asyncio.to_thread(
                self._read_excel_with_header, file_path
            )

            logger.info(
                f"Loaded {len(excel_df)} строк. "
                f"Header in row: {header_row + 1}"
            )

            # 2. Объединение данных
            logger.info("Merging with supplier data")
            merged_df = await asyncio.to_thread(
                self._merge_with_supplier_data, excel_df, supplier_data
            )

            # 3. Заполнение пропусков
            logger.info("Filling missing data")
            filled_df = await asyncio.to_thread(
                self._fill_missing_data, merged_df
            )

            # 4. Применение правил обработки
            logger.info("Applying processing rules")
            processed_df = await asyncio.to_thread(
                self._apply_processing_rules, filled_df
            )

            # 5. Сохранение результата
            output_file = self._generate_output_filename(file_path)
            logger.info("Saving result", extra={"path": str(output_file)})

            await asyncio.to_thread(
                self._save_to_excel_with_formatting,
                file_path,
                output_file,
                processed_df,
                header_row,
            )

            # 6. Логирование статистики
            self._log_processing_statistics(processed_df)

        except Exception as e:
            logger.error(
                "Excel processing failed",
                extra={"error": str(e)},
            )
            raise ExcelProcessingError(
                error_code="EXCEL_PROCESSING_ERROR",
                message="Failed to process Excel file",
                details=str(e),
            ) from e
        else:
            return output_file

    # ----- Чтение Excel -----

    def _read_excel_with_header(
        self, file_path: Path
    ) -> tuple[pd.DataFrame, int]:
        """
        Читает Excel файл, автоматически определяя строку с заголовками.

        Returns:
            Tuple[DataFrame, int]: DataFrame и номер строки заголовка
        """
        try:
            # Читаем первые строки для анализа
            preview_df = pd.read_excel(file_path, header=None, nrows=20)

            # Ищем строку с заголовком "Код"
            header_row = None
            for i in range(len(preview_df)):
                if "Код" in preview_df.iloc[i].astype(str).values:
                    header_row = i
                    logger.debug(f"The title was found in the line {i}")
                    break

            self._validate_header_row(header_row)

            # Читаем с правильным заголовком
            df = pd.read_excel(file_path, header=header_row)

            # Стандартизация названий колонок
            column_mapping = self._standardize_column_names(df.columns)
            df = df.rename(columns=column_mapping)

            # Проверяем обязательные колонки
            required_columns = ["Код", "Наименование", "Цена"]
            missing_columns = [
                col for col in required_columns if col not in df.columns
            ]

            self._validate_missing_columns(missing_columns)

            # Очистка данных
            df["Код"] = df["Код"].astype(str).str.strip()

        except BaseAppException:
            raise

        except Exception as e:
            logger.error(f"Ошибка при чтении Excel: {e}")
            raise ExcelProcessingError(
                message="Ошибка при чтении Excel файла", details=str(e)
            ) from e
        else:
            return df, header_row  # type: ignore[return-value]

    def _validate_header_row(self, header_row: int | None) -> None:
        if header_row is None:
            raise ExcelProcessingError(
                error_code="HEADER_NOT_FOUND",
                message="Could not find header row containing 'Код'",
            )

    def _validate_missing_columns(self, missing_columns: list[str]) -> None:
        if missing_columns:
            raise ExcelProcessingError(
                error_code="MISSING_COLUMNS",
                message=f"Missing required columns: {missing_columns}",
            )

    def _standardize_column_names(self, columns: pd.Index) -> dict[str, str]:
        """
        Стандартизирует названия колонок.

        Args:
            columns: Исходные названия колонок

        Returns:
            Dict: Словарь для переименования
        """
        mapping: dict[str, str] = {}

        for col in columns:
            col_str = str(col).lower()

            if "код" in col_str:
                mapping[col] = "Код"
            elif "наимен" in col_str or "name" in col_str:
                mapping[col] = "Наименование"
            elif "цена" in col_str or "price" in col_str:
                mapping[col] = "Цена"
            elif "заказ" in col_str or "order" in col_str:
                mapping[col] = "заказ"
            elif "сумм" in col_str or "sum" in col_str:
                mapping[col] = "Сумма"
            else:
                mapping[col] = col  # Оставляем как есть

        return mapping

    # ----- Слияние -----

    def _merge_with_supplier_data(
        self, excel_df: pd.DataFrame, supplier_data: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Объединяет данные Excel с данными поставщика.

        Returns:
            DataFrame: Объединенные данные
        """
        try:
            excel_df["Код"] = pd.to_numeric(excel_df["Код"], errors="coerce")

            merged_df = pd.merge(
                excel_df,
                supplier_data[["code", "category", "subcategory"]],
                left_on="Код",
                right_on="code",
                how="left",
            )

            # Удаляем вспомогательную колонку
            merged_df = merged_df.drop(columns=["code"])

            # Статистика совпадений
            matches = merged_df["category"].notna().sum()
            match_percentage = (matches / len(merged_df)) * 100

            logger.info(
                "Merge statistics",
                extra={
                    "matches": matches,
                    "total": len(merged_df),
                    "percent": match_percentage,
                },
            )

        except Exception as e:
            raise PriceProcessingError(
                error_code="MERGE_SUPPLIER_DATA_WITH_EXCEL_ERROR",
                message="Error when combining data",
                details=str(e),
            ) from e
        else:
            return merged_df

    # ----- Заполнение пропусков -----

    def _fill_missing_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Заполняет пропущенные значения различными методами.

        Returns:
            DataFrame: Данные с заполненными пропусками
        """
        result_df = df.copy()

        # Добавляем колонки для отслеживания заполнения
        result_df["_filled_by_rule"] = False
        result_df["_filled_by_neighbors"] = False

        # 1. Заполнение на основе соседних строк
        # result_df = self._fill_by_neighbors(result_df)
        result_df = self._fill_missing_from_neighbors(result_df)

        # 2. Заполнение на основе наименования товара
        result_df = self._fill_by_product_name(result_df)

        return result_df

    def _fill_missing_from_neighbors(
        self,
        df: pd.DataFrame,
        group_col: str = "category",
        subgroup_col: str = "subcategory",
    ) -> pd.DataFrame:
        """
        Заполняет пропущенные значения на основе совпадающих значений выше и
        ниже. Помечает строки, которые были заполнены.

        Args:
            df: Исходный DataFrame
            group_col: Название колонки с группой товаров
            subgroup_col: Название колонки с подгруппой

        Returns:
            DataFrame с заполненными значениями и пометками
        """
        # Создаем копию чтобы не изменять оригинал
        result_df = df.copy()

        # Проходим по всем строкам
        for i in range(len(result_df)):
            # Проверяем, что текущая строка имеет пропуски
            current_group = result_df.at[i, group_col]
            current_subgroup = result_df.at[i, subgroup_col]

            if pd.isna(current_group) or pd.isna(current_subgroup):
                # Ищем ближайшие заполненные строки выше
                upper_group = None
                upper_subgroup = None
                for j in range(i - 1, -1, -1):
                    if not pd.isna(
                        result_df.at[j, group_col]
                    ) and not pd.isna(result_df.at[j, subgroup_col]):
                        upper_group = result_df.at[j, group_col]
                        upper_subgroup = result_df.at[j, subgroup_col]
                        break

                # Ищем ближайшие заполненные строки ниже
                lower_group = None
                lower_subgroup = None
                for j in range(i + 1, len(result_df)):
                    if not pd.isna(
                        result_df.at[j, group_col]
                    ) and not pd.isna(result_df.at[j, subgroup_col]):
                        lower_group = result_df.at[j, group_col]
                        lower_subgroup = result_df.at[j, subgroup_col]
                        break

                # Если значения сверху и снизу совпадают и не None
                if (
                    upper_group is not None
                    and lower_group is not None
                    and upper_group == lower_group
                    and upper_subgroup == lower_subgroup
                ):
                    # Заполняем пропуски
                    if pd.isna(current_group):
                        result_df.at[i, group_col] = upper_group
                        result_df.at[i, "_filled_by_neighbors"] = True
                        # result_df.at[i, '_group_filled'] = True

                    if pd.isna(current_subgroup):
                        result_df.at[i, subgroup_col] = upper_subgroup
                        result_df.at[i, "_filled_by_neighbors"] = True
                        # result_df.at[i, '_subgroup_filled'] = True

        return result_df

    def _fill_by_product_name(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Заполняет пропуски на основе наименования товара.
        """
        result_df = df.copy()

        for idx, row in result_df.iterrows():
            product_name = row.get("Наименование", "")

            if pd.notna(product_name):
                for condition, group, subgroup in self.PRODUCT_NAME_RULES:
                    if condition(product_name):
                        result_df.at[idx, "category"] = group
                        result_df.at[idx, "subcategory"] = subgroup
                        result_df.at[idx, "_filled_by_rule"] = True
                        break

        return result_df

    # def _fill_from_product_name(self, df: pd.DataFrame) -> pd.DataFrame:
    #     def _update_row(row: pd.Series) -> pd.Series:
    #         name = row["Наименование"]
    #         if any(
    #             keyword in name.lower() for keyword in ["лицензия", "***"]
    #         ):
    #             row["category"] = "Элит Парфюм"
    #             row["subcategory"] = "Лицензия***"
    #         elif any(
    #             keyword in name.lower()
    #             for keyword in [
    #                 "vintag",
    #                 "винтаж",
    #                 "novaya zarya",
    #                 "косметика",
    #             ]
    #         ) or ("MONTALE" in name and "декодированный" in name):
    #             row["category"] = "NO"
    #             row["subcategory"] = "NO"
    #         return row

    #     return df.apply(_update_row, axis=1)

    # def _fill_by_neighbors(self, df: pd.DataFrame) -> pd.DataFrame:
    #     """
    #     Заполняет пропуски на основе значений в соседних строках.
    #     Улучшенная версия: заполняет только если все строки в блоке
    #     одинаковы.
    #     """
    #     result_df = df.copy()

    #     # Векторизованный подход для производительности
    #     forward_filled = result_df[["category", "subcategory"]].ffill()
    #     backward_filled = result_df[["category", "subcategory"]].bfill()

    #     # Маска строк, где значения вперед и назад совпадают
    #     mask = (
    #         (forward_filled["category"] == backward_filled["category"])
    #         & (
    #             forward_filled["subcategory"]
    #             == backward_filled["subcategory"]
    #         )
    #         & (
    #                result_df["category"].isna()
    #                | result_df["subcategory"].isna()
    #           )
    #     )

    #     # Заполняем совпадающие строки
    #     result_df.loc[mask, "category"] = forward_filled.loc[
    #         mask, "category"
    #     ]
    #     result_df.loc[mask, "subcategory"] = forward_filled.loc[
    #         mask, "subcategory"
    #     ]
    #     result_df.loc[mask, "_filled_by_neighbors"] = True

    #     filled_count = mask.sum()
    #     if filled_count > 0:
    #         logger.info(
    #             f"Заполнено {filled_count} строк на основе соседних "
    #             "значений"
    #         )

    #     return result_df

    # ----- Применение правил -----

    def _apply_processing_rules(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Применяет финальные правила обработки к данным.
        """
        result_df = df.copy()

        # Удаляем временные колонки если нужно
        if "_filled_by_rule" in result_df.columns:
            # Можно сохранить статистику перед удалением
            filled_by_rule = result_df["_filled_by_rule"].sum()
            filled_by_neighbors = result_df["_filled_by_neighbors"].sum()

            logger.debug(
                f"Field by rules: {filled_by_rule}, "
                f"neighbors: {filled_by_neighbors}"
            )

            # Удаляем временные колонки
            result_df = result_df.drop(
                columns=["_filled_by_rule", "_filled_by_neighbors"]
            )

        return result_df

    # ----- Генерация имени файла -----

    def _generate_output_filename(self, input_path: Path) -> Path:
        """
        Генерирует имя для выходного файла.

        Args:
            input_path: Путь к исходному файлу

        Returns:
            Path: Путь для сохранения результата
        """
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        input_stem = input_path.stem

        output_name = f"{input_stem}_enriched_{timestamp}.xlsx"
        output_path = input_path.parent / output_name

        return output_path

    # ----- Сохранение Excel -----

    def _save_to_excel_with_formatting(
        self,
        original_file: Path,
        output_file: Path,
        df: pd.DataFrame,
        header_row: int,
    ) -> None:
        """
        Сохраняет DataFrame в Excel с сохранением форматирования оригинала.
        """
        try:
            # Загружаем оригинальную книгу
            wb = load_workbook(original_file)
            ws = wb.active

            # Находим позицию для вставки новых колонок
            # (после колонки "Сумма" или последней существующей)
            last_column = ws.max_column

            # Вставляем 2 новые колонки
            ws.insert_cols(last_column + 1, 2)

            # Записываем заголовки новых колонок
            group_col_letter = get_column_letter(last_column + 1)
            subgroup_col_letter = get_column_letter(last_column + 2)

            ws[f"{group_col_letter}{header_row + 1}"] = "Группа товара"
            ws[f"{subgroup_col_letter}{header_row + 1}"] = "Подгруппа"

            # Записываем данные
            for idx, row in df.iterrows():
                excel_row = header_row + 2 + idx

                # Записываем группу
                group_value = row.get("category")
                if pd.notna(group_value):
                    ws[f"{group_col_letter}{excel_row}"] = group_value

                # Записываем подгруппу
                subgroup_value = row.get("subcategory")
                if pd.notna(subgroup_value):
                    ws[f"{subgroup_col_letter}{excel_row}"] = subgroup_value

            # Сохраняем
            wb.save(output_file)
            logger.info(f"Файл успешно сохранен: {output_file}")

        except Exception as e:
            logger.error(f"Ошибка при сохранении Excel файла: {e}")
            raise

    # def _write_to_excel_with_formatting(
    #     self,
    #     original_file: str,
    #     new_file: str,
    #     df_merged: pd.DataFrame,
    #     header_row: int,
    # ) -> None:
    #     """
    #     Записывает данные обратно в Excel с сохранением оригинального
    #     форматирования
    #     """

    #     # Загружаем оригинальную книгу
    #     wb = load_workbook(original_file)
    #     ws = wb.active

    #     # Определяем колонки для новых данных
    #     category_col = get_column_letter(
    #         df_merged.columns.get_loc("category") + 1
    #     )
    #     subgroup_col = get_column_letter(
    #         df_merged.columns.get_loc("subcategory") + 1
    #     )

    #     # Записываем заголовки новых колонок
    #     ws[f"{category_col}{header_row + 1}"] = "Группа товара"
    #     ws[f"{subgroup_col}{header_row + 1}"] = "Подгруппа"

    #     # Записываем данные
    #     for idx, row in df_merged.iterrows():
    #         excel_row = (
    #             header_row + 2 + idx
    #         )  # +2 потому что заголовок на header_row+1

    #         # Записываем category
    #         category = row["category"]
    #         if pd.notna(category):
    #             ws[f"{category_col}{excel_row}"] = category

    #         # Записываем subgroup
    #         subgroup = row["subcategory"]
    #         if pd.notna(subgroup):
    #             ws[f"{subgroup_col}{excel_row}"] = subgroup

    #     # Сохраняем в новый файл
    #     wb.save(new_file)
    #     print(f"Файл сохранен: {new_file}")

    # ----- Статистика -----

    def _log_processing_statistics(self, df: pd.DataFrame) -> None:
        """
        Логирует статистику обработки данных.

        Args:
            df: Обработанный DataFrame
        """
        try:
            total_rows = len(df)
            group_missing = df["category"].isna().sum()
            subgroup_missing = df["subcategory"].isna().sum()

            # group_fill_rate = (
            #     (total_rows - group_missing) / total_rows
            # ) * 100
            # subgroup_fill_rate = (
            #     (total_rows - subgroup_missing) / total_rows
            # ) * 100

            logger.info(
                "Processing statistics",
                extra={
                    "total_rows": total_rows,
                    "category_filled": total_rows - group_missing,
                    "category_missing": group_missing,
                    "subcategory_filled": total_rows - subgroup_missing,
                    "subcategory_missing": subgroup_missing,
                },
            )

        except (KeyError, AttributeError, TypeError, ZeroDivisionError) as e:
            logger.warning(f"Couldn't collect statistics: {e}")

    # ----- Очистка -----

    async def _cleanup_temp_files(self, file_paths: list[Path]) -> None:
        """
        Асинхронно удаляет временные файлы.

        Args:
            file_paths: Список путей к файлам для удаления
        """
        for file_path in file_paths:
            if file_path and file_path.exists():
                try:
                    # Асинхронное удаление файла
                    await asyncio.to_thread(os.remove, file_path)
                    logger.debug(
                        "Removed temp file", extra={"path": str(file_path)}
                    )
                except FileNotFoundError:
                    logger.debug(
                        f"The file has already been deleted: {file_path}"
                    )
                except PermissionError as e:
                    logger.warning(
                        "Don't have the rights to delete the file. "
                        f"{file_path}: {e}"
                    )
                except OSError as e:
                    # Ловим все OS-специфичные ошибки
                    logger.warning(
                        "Failed to remove temp file",
                        extra={"path": str(file_path), "error": str(e)},
                    )
                except RuntimeError as e:
                    # Ошибки, связанные с asyncio.to_thread
                    logger.warning(
                        "Execution error when deleting a file "
                        f"{file_path}: {e}"
                    )


# ===== Dependency Injection =====


def get_price_loader(
    settings: Annotated[Any, Depends(lambda: settings)],
    supplier_codes_repo: Annotated[
        SupplierProductCodeRepository,
        Depends(get_supplier_product_codes_repo),
    ],
    file_uploader: Annotated[FileUploader, Depends(get_file_uploader)],
    supplier_id: int = PriceLoader.DEFAULT_SUPPLIER_ID,
) -> PriceLoader:
    """
    Dependency для получения экземпляра PriceLoader.

    Args:
        settings: Настройки приложения
        supplier_codes_repo: Репозиторий данных поставщика
        supplier_id: ID поставщика

    Returns:
        PriceLoader: Экземпляр сервиса
    """
    return PriceLoader(
        settings=settings,
        supplier_codes_repo=supplier_codes_repo,
        file_uploader=file_uploader,
        supplier_id=supplier_id,
    )
